import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


class Vacancy:
    currenciesis = {"AZN": 35.68,
                  "BYR": 23.91,
                  "EUR": 59.90,
                  "GEL": 21.74,
                  "KGS": 0.76,
                  "KZT": 0.13,
                  "RUR": 1,
                  "UAH": 1.64,
                  "USD": 60.66,
                  "UZS": 0.0055, }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currenciesis[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])




class DataSet:

    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name


    @staticmethod
    def average(dictionary):
        new_dictionary = {}
        for key, values in dictionary.items():
            new_dictionary[key] = int(sum(values) / len(values))
        return new_dictionary


    @staticmethod
    def increment(dictionary, key, amount):
        if key in dictionary:
            dictionary[key] += amount
        else:
            dictionary[key] = amount


    def get_statistic(self):
        salary = {}
        salary_name_of_vacancy = {}
        salary_of_city = {}
        number_of_vacancies = 0

        for vacancy_dictionary in self.csv_reader():
            vacancy = Vacancy(vacancy_dictionary)
            self.increment(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(salary_name_of_vacancy, vacancy.year, [vacancy.salary_average])
            self.increment(salary_of_city, vacancy.area_name, [vacancy.salary_average])
            number_of_vacancies += 1

        vacancies_number = dict([(key, len(value)) for key, value in salary.items()])
        vacancies_number_by_name = dict([(key, len(value)) for key, value in salary_name_of_vacancy.items()])

        if not salary_name_of_vacancy:
            salary_name_of_vacancy = dict([(key, [0]) for key, value in salary.items()])
            vacancies_number_by_name = dict([(key, 0) for key, value in vacancies_number.items()])

        statistics1 = self.average(salary)
        statistics2 = self.average(salary_name_of_vacancy)
        statistics3 = self.average(salary_of_city)

        statistics4 = {}
        for year, salaries in salary_of_city.items():
            statistics4[year] = round(len(salaries) / number_of_vacancies, 4)
        statistics4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in statistics4.items()]))
        statistics4.sort(key=lambda a: a[-1], reverse=True)
        statistics5 = statistics4.copy()
        statistics4 = dict(statistics4)
        statistics3 = list(filter(lambda a: a[0] in list(statistics4.keys()), [(key, value) for key, value in statistics3.items()]))
        statistics3.sort(key=lambda a: a[-1], reverse=True)
        statistics3 = dict(statistics3[:10])
        statistics5 = dict(statistics5[:10])

        return statistics1, vacancies_number, statistics2, vacancies_number_by_name, statistics3, statistics5

    def csv_reader(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    @staticmethod
    def print_statistics(statistics1, statistics2, statistics3, statistics4, statistics5, statistics6):
        print('Динамика уровня зарплат по годам: {0}'.format(statistics1))
        print('Динамика количества вакансий по годам: {0}'.format(statistics2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statistics3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(statistics4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(statistics5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(statistics6))



class InputConnect:

    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        statistics1, statistics2, statistics3, statistics4, statistics5, statistics6 = dataset.get_statistic()
        dataset.print_statistics(statistics1, statistics2, statistics3, statistics4, statistics5, statistics6)

        report = Report(self.vacancy_name, statistics1, statistics2, statistics3, statistics4, statistics5, statistics6)
        report.generate_excel()




class Report:

    def __init__(self, vacancy_name, statistics1, statistics2, statistics3, statistics4, statistics5, statistics6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.statistics1 = statistics1
        self.statistics2 = statistics2
        self.statistics3 = statistics3
        self.statistics4 = statistics4
        self.statistics5 = statistics5
        self.statistics6 = statistics6

    def generate_excel(self):
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.statistics1.keys():
            ws1.append([year, self.statistics1[year], self.statistics3[year], self.statistics2[year], self.statistics4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column_width = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_width) > i:
                    if len(cell) > column_width[i]:
                        column_width[i] = len(cell)
                else:
                    column_width += [len(cell)]

        for i, column_width in enumerate(column_width, 1):  # ,1 to start at 1
            ws1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.statistics5.items(), self.statistics6.items()):
            data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            ws2.append(row)

        column_width = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_width) > i:
                    if len(cell) > column_width[i]:
                        column_width[i] = len(cell)
                else:
                    column_width += [len(cell)]

        for i, column_width in enumerate(column_width, 1):  # ,1 to start at 1
            ws2.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            ws1[col + '1'].font = font_bold
            ws2[col + '1'].font = font_bold

        for index, _ in enumerate(self.statistics5):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                ws2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.statistics1[1] = 1
        for row, _ in enumerate(self.statistics1):
            for col in 'ABCDE':
                ws1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.wb.save('report.xlsx')


if __name__ == '__main__':
    InputConnect()